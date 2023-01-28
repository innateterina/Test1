#Прочитать сохранённый csv-файл из задания №19 и сохранить данные в excel-файл,
# кроме возраста – столбец с этими данными не нужен.

import csv
with open("voc_file02.csv", encoding='utf-8') as r_file:
    file_reader = csv.reader(r_file)
    count = 0
    data = []
    for item in file_reader:
        if count == 0:
            name_of_columns = item
        else:
            data.append(item)
        count += 1

print(name_of_columns)
print(data)

import openpyxl

wb = openpyxl.Workbook()
wb.create_sheet(title='First sheet', index=0)
sheet = wb['First sheet']
persons = ['person_1', 'person_2', 'person_3', 'person_4', 'person_5']
for r, a in enumerate(persons):
    sheet.cell(row=1, column=r+2).value = a
for col, b in enumerate(name_of_columns):
    sheet.cell(row=col+2, column=1).value = b

field = data[0]
field_1 = data[1]
field_2 = data[2]
field_3 = data[3]
field_4 = data[4]

for colum_index, colum in enumerate((field, field_1, field_2, field_3, field_4)):
    for row_index, value in enumerate(colum):
        cell = sheet.cell(row=row_index+2, column=colum_index+2)
        cell.value = value

import random
num_list = []
while len(num_list) < 5:
    num = str(random.randint(0, 999999999)).zfill(9)
    if not num in num_list:
        num_list.append(num)
for m, a in enumerate(num_list):
    sheet.cell(row=5, column=m+2).value = a
sheet. delete_rows(4)
wb.save('voc_file03.xlsx')
